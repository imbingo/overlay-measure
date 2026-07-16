from __future__ import annotations

import numpy as np
import pytest
from openpyxl import load_workbook

from overlay_measure.auto_mark_detector import AutoDetectReport, detect_auto_marks_with_report
from overlay_measure.image_loader import load_image
from overlay_measure.measurement_units import radial_diameter_residual_um, rotated_rect_size_um
from overlay_measure.models import DetectionParams, MeasurementConfig, DetectionResult, OverlayResult
from overlay_measure.overlay_calculator import calculate_relative_overlay
from overlay_measure.result_exporter import export_results, resize_dimensions_preserving_aspect
from overlay_measure.rz_calculator import build_summary_rows


def test_non_square_pixel_radial_distances_use_xy_scales():
    config = MeasurementConfig(pixel_size_x_um=0.1, pixel_size_y_um=0.2)
    points = np.asarray([[10.0, 0.0], [0.0, 10.0]], dtype=np.float64)
    diameter_um, residual_um = radial_diameter_residual_um(points, 0.0, 0.0, 10.0, 0.0, config)

    assert diameter_um == pytest.approx(3.0)
    assert residual_um == pytest.approx(0.5)


def test_rotated_rectangle_size_uses_axis_angle():
    config = MeasurementConfig(pixel_size_x_um=0.1, pixel_size_y_um=0.2)

    width_um, height_um = rotated_rect_size_um(10.0, 20.0, 0.0, config)
    assert width_um == pytest.approx(1.0)
    assert height_um == pytest.approx(4.0)

    width_um, height_um = rotated_rect_size_um(10.0, 20.0, 90.0, config)
    assert width_um == pytest.approx(2.0)
    assert height_um == pytest.approx(2.0)


def test_auto_detect_report_exposes_limits():
    image = load_image("sample_data/sample_square_single.png")
    report = detect_auto_marks_with_report(
        image.gray,
        "upper",
        DetectionParams(canny_low=25, canny_high=90, min_gradient=2, residual_limit_px=0.6),
        0.1,
        0.1,
    )

    assert len(report.results) >= 1
    assert report.total_contours >= report.processed_contours
    assert report.max_results == 48


def test_auto_detect_report_warning_text():
    report = AutoDetectReport(
        results=[],
        processed_contours=0,
        selected_contours=96,
        total_contours=400,
        max_contours_per_mask=96,
        max_results=48,
        time_limit_s=4.0,
        truncated_by_time=True,
        truncated_by_contour_limit=True,
        truncated_by_result_limit=True,
    )

    text = report.warning_text()
    assert "时间上限" in text
    assert "轮廓过多" in text
    assert "候选达到上限" in text


def test_rz_summary_is_testable_outside_ui():
    config = MeasurementConfig(rz_layout="Y向前后分布", rz_distance_l_um=1000.0, rz_limit=600.0)
    rows = build_summary_rows(
        {
            "Mark1": OverlayResult("Mark1", 0, 0, 1.0, 0.0, 1.0, "Pass"),
            "Mark2": OverlayResult("Mark2", 0, 0, 1.5, 0.0, 1.5, "Pass"),
        },
        config,
    )

    rz_row = rows[-1]
    assert rz_row["项目"] == "Rz"
    assert rz_row["Rz(μrad)"] == pytest.approx(500.0)
    assert rz_row["判定"] == "通过"


def test_angle_thickness_compensation_uses_correct_axes():
    config = MeasurementConfig(
        pixel_size_x_um=0.1,
        pixel_size_y_um=0.1,
        material_thickness_mm=2.0,
        rx_angle_urad=100.0,
        ry_angle_urad=50.0,
    )
    reference = DetectionResult("Mark1", "upper", 10, 20, 1, 2, 1, 0.1, 0, 0, 10, 1, "Circle")
    target = DetectionResult("Mark1", "upper", 30, 10, 3, 1, 1, 0.1, 0, 0, 10, 1, "Circle")

    overlay = calculate_relative_overlay("Mark1", reference, target, config)

    assert overlay.delta_x_um == pytest.approx(2.1)
    assert overlay.delta_y_um == pytest.approx(0.8)


def test_export_repeatability_sheet_and_mark_image_aspect_ratio(tmp_path):
    image_path = tmp_path / "mark.png"
    from PIL import Image

    Image.new("L", (400, 100), color=128).save(image_path)
    out_path = tmp_path / "result.xlsx"

    export_results(
        str(out_path),
        rows=[],
        config=MeasurementConfig(),
        summary_rows=[{"项目": "Mark1", "Dx1(μm)": 1.23456, "判定": "通过"}],
        repeatability_rows=[
            {"Mark": "Mark1", "次数": 1, "Dx(μm)": 1.0, "PV-Dx(μm)": 0.2},
            {"Mark": "Mark1", "次数": "统计", "Dx(μm)": 1.1, "PV-Dx(μm)": 0.2},
        ],
        mark_images=[{"mark_id": "Mark1", "layer": "上层", "path": str(image_path), "note": ""}],
    )

    wb = load_workbook(out_path)
    assert "多次测量结果" in wb.sheetnames
    assert len(wb["Mark图片"]._images) == 1
    assert resize_dimensions_preserving_aspect(400, 100, 260, 180) == (260, 65)
