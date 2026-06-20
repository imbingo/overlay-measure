from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .measurement_units import axis_scale_um_per_px, mean_pixel_size_um, rotated_rect_size_um, scalar_px_to_um
from .models import DetectionResult, MeasurementConfig, OverlayResult


DETAIL_COLUMNS = {
    "timestamp": "测量时间",
    "measurement_mode": "测量模式",
    "upper_file": "上层/单图文件",
    "lower_file": "下层文件",
    "pixel_size_x_um": "像素尺寸X(μm/px)",
    "pixel_size_y_um": "像素尺寸Y(μm/px)",
    "registration_offset_x_um": "配准偏移X(μm)",
    "registration_offset_y_um": "配准偏移Y(μm)",
    "mark_id": "标记",
    "layer": "层",
    "center_x_um": "中心X(μm)",
    "center_y_um": "中心Y(μm)",
    "diameter_um": "直径/尺寸(μm)",
    "fit_residual_um": "参考残差(μm)",
    "edge_point_count": "边缘点数",
    "confidence": "置信度",
    "fitting_mode": "拟合模式",
    "measurement_stage": "测量阶段",
    "quality_status": "质量状态",
    "coverage": "覆盖率",
    "rejected_count": "剔除点数",
    "rejected_ratio": "异常点比例",
    "max_deviation_um": "最大轮廓偏差(μm)",
    "failure_reason": "失效原因",
    "recipe_validation_status": "配方验证状态",
    "detection_warning": "识别提示",
    "shape_width_um": "宽度(μm)",
    "shape_height_um": "高度(μm)",
    "shape_major_um": "长轴(μm)",
    "shape_minor_um": "短轴(μm)",
    "shape_angle_deg": "角度(°)",
    "shape_aspect_ratio": "宽高比",
    "roi_type": "ROI类型",
    "roi_inner_ratio": "内环比例",
    "roi_inner_radius_um": "ROI内半径(μm)",
    "roi_outer_radius_um": "ROI外半径(μm)",
    "caliper_count": "卡尺数量",
    "caliper_width_um": "卡尺宽度(μm)",
    "search_direction": "搜索方向",
    "roi_target_edge": "边缘选择",
    "roi_angle_deg": "ROI角度(°)",
    "delta_x_um": "ΔX(μm)",
    "delta_y_um": "ΔY(μm)",
    "overlay_r_um": "Dxy/R(μm)",
    "result": "结果",
    "overlay_warning": "对位提示",
}


def _mode_cn(mode: str) -> str:
    return "双图模式" if mode == "Dual Image" else "单图模式"


def _layer_cn(layer: str) -> str:
    return {"upper": "上层", "lower": "下层"}.get(layer, layer)


def _result_cn(result: str) -> str:
    return {"Pass": "通过", "Fail": "不通过", "Trial": "试测/不判定"}.get(result, result)


def _fit_cn(mode: str) -> str:
    return {
        "EdgeCenter": "边缘中心",
        "RegionCenter": "区域中心",
        "CaliperCircle": "卡尺找圆",
        "AutoCircle": "自动圆轮廓",
        "AutoRectangle": "自动方形轮廓",
        "ProductionCircle": "正式卡尺圆拟合",
        "ProductionRectangle": "正式四边卡尺拟合",
        "Auto": "自动",
        "Circle": "圆",
        "Ellipse": "椭圆",
        "Rectangle": "矩形/方孔",
    }.get(mode, mode)


def _roi_cn(value: str) -> str:
    return {
        "Annulus": "圆环",
        "Caliper Circle": "卡尺圆",
        "Rectangular Ring": "矩形环",
        "Circle": "圆",
        "Rectangle": "矩形",
        "Auto Full Image": "全图自动识别",
        "Auto Caliper Circle": "自动卡尺圆",
        "Auto Four-Side Caliper": "自动四边卡尺",
        "All Edges": "全部边缘",
        "Near Inner Boundary": "靠近内环",
        "Near Outer Boundary": "靠近外环",
        "Strongest Edge": "最强边缘",
    }.get(value, value)


def _direction_cn(value: str) -> str:
    return {
        "Inner to Outer": "由内向外",
        "Outer to Inner": "由外向内",
    }.get(value, value)


def build_detection_rows(
    detections: Dict[str, Dict[str, DetectionResult]],
    overlays: Dict[str, OverlayResult],
    config: MeasurementConfig,
    upper_file: str = "",
    lower_file: str = "",
) -> List[dict]:
    rows = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mean_scale_um = mean_pixel_size_um(config)
    for mark_id, layer_map in detections.items():
        overlay = overlays.get(mark_id)
        for layer, det in layer_map.items():
            width_px = det.shape_params.get("width_px")
            height_px = det.shape_params.get("height_px")
            angle_deg = float(det.shape_params.get("angle_deg", 0.0))
            width_um = height_um = None
            if width_px is not None and height_px is not None:
                width_um, height_um = rotated_rect_size_um(float(width_px), float(height_px), angle_deg, config)
            major_px = det.shape_params.get("major_px")
            minor_px = det.shape_params.get("minor_px")
            row = {
                "timestamp": now,
                "measurement_mode": _mode_cn(config.mode),
                "upper_file": upper_file,
                "lower_file": lower_file,
                "pixel_size_x_um": config.pixel_size_x_um,
                "pixel_size_y_um": config.pixel_size_y_um,
                "registration_offset_x_um": config.registration_offset_x_um,
                "registration_offset_y_um": config.registration_offset_y_um,
                "mark_id": mark_id,
                "layer": _layer_cn(layer),
                "center_x_um": det.center_x_um,
                "center_y_um": det.center_y_um,
                "diameter_um": det.diameter_um,
                "fit_residual_um": det.residual_um,
                "edge_point_count": det.edge_point_count,
                "confidence": det.confidence,
                "fitting_mode": _fit_cn(det.fitting_mode),
                "measurement_stage": "正式精测" if det.shape_params.get("measurement_stage") == "production_measurement" else "候选检测",
                "quality_status": {"Valid": "有效", "Invalid": "无效"}.get(det.shape_params.get("quality_status"), det.shape_params.get("quality_status", "")),
                "coverage": det.shape_params.get("coverage"),
                "rejected_count": det.shape_params.get("rejected_count"),
                "rejected_ratio": det.shape_params.get("rejected_ratio"),
                "max_deviation_um": det.shape_params.get("max_deviation_um"),
                "failure_reason": det.shape_params.get("failure_reason"),
                "recipe_validation_status": {"Draft": "草稿/未验证", "Validated": "已验证/正式生产"}.get(det.shape_params.get("recipe_validation_status"), det.shape_params.get("recipe_validation_status", "")),
                "detection_warning": det.warning,
                "shape_width_um": width_um,
                "shape_height_um": height_um,
                "shape_major_um": major_px * axis_scale_um_per_px(config, angle_deg) if major_px is not None else None,
                "shape_minor_um": minor_px * axis_scale_um_per_px(config, angle_deg + 90.0) if minor_px is not None else None,
                "shape_angle_deg": det.shape_params.get("angle_deg"),
                "shape_aspect_ratio": det.shape_params.get("aspect_ratio"),
                "roi_type": _roi_cn(det.shape_params.get("roi_type")),
                "roi_inner_ratio": det.shape_params.get("roi_inner_ratio"),
                "roi_inner_radius_um": scalar_px_to_um(det.shape_params.get("roi_inner_radius_px"), config) if det.shape_params.get("roi_inner_radius_px") is not None else None,
                "roi_outer_radius_um": scalar_px_to_um(det.shape_params.get("roi_outer_radius_px"), config) if det.shape_params.get("roi_outer_radius_px") is not None else None,
                "caliper_count": det.shape_params.get("caliper_count"),
                "caliper_width_um": scalar_px_to_um(det.shape_params.get("caliper_width_px"), config) if det.shape_params.get("caliper_width_px") is not None else None,
                "search_direction": _direction_cn(det.shape_params.get("search_direction")),
                "roi_target_edge": _roi_cn(det.shape_params.get("roi_target_edge")),
                "roi_angle_deg": det.shape_params.get("roi_angle_deg"),
                "delta_x_um": overlay.delta_x_um if overlay else None,
                "delta_y_um": overlay.delta_y_um if overlay else None,
                "overlay_r_um": overlay.overlay_r_um if overlay else None,
                "result": _result_cn(overlay.result) if overlay else "",
                "overlay_warning": overlay.warning if overlay else "",
            }
            rows.append(row)
    return rows


def _autosize(ws):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_len = 8
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, min(36, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def _style_sheet(ws, fail_columns: Optional[List[str]] = None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    fail_font = Font(color="9C0006", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    fail_columns = fail_columns or []
    headers = [cell.value for cell in ws[1]]
    fail_indexes = [headers.index(col) + 1 for col in fail_columns if col in headers]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        for idx in fail_indexes:
            value = row[idx - 1].value
            if value in {"失败", "NG"} or (isinstance(value, str) and "超限" in value):
                row[idx - 1].fill = fail_fill
                row[idx - 1].font = fail_font
    _autosize(ws)


def export_results(
    path: str,
    rows: List[dict],
    config: Optional[MeasurementConfig] = None,
    summary_rows: Optional[List[dict]] = None,
    mark_images: Optional[List[dict]] = None,
) -> None:
    detail_df = pd.DataFrame(rows).rename(columns=DETAIL_COLUMNS)
    summary_df = pd.DataFrame(summary_rows or [])
    # Keep exported measurement results concise for production review.
    # Internal calculation remains full precision; only output tables are rounded.
    detail_df = detail_df.round(3)
    summary_df = summary_df.round(3)
    ext = Path(path).suffix.lower()
    if ext != ".xlsx":
        # CSV can contain only one table, so export the concise summary when available.
        df = summary_df if not summary_df.empty else detail_df
        df.to_csv(path, index=False, encoding="utf-8-sig")
        return

    info_rows = []
    if config is not None:
        info_rows = [
            {"项目": "物料编码", "内容": config.material_code},
            {"项目": "配方名称", "内容": getattr(config, "recipe_name", "")},
            {"项目": "配方版本", "内容": getattr(config, "recipe_version", "")},
            {"项目": "配方验证状态", "内容": "已验证/正式生产" if getattr(config, "recipe_validation_status", "Draft") == "Validated" else "草稿/未验证"},
            {"项目": "结果用途", "内容": "正式判定" if getattr(config, "recipe_validation_status", "Draft") == "Validated" else "试测/未验证配方，不作正式判定"},
            {"项目": "工序", "内容": config.process_name},
            {"项目": "测量设备型号", "内容": config.equipment_model},
            {"项目": "设备校准日期", "内容": config.calibration_date},
            {"项目": "操作人员", "内容": config.operator_name},
            {"项目": "测量模式", "内容": _mode_cn(config.mode)},
            {"项目": "工作方式", "内容": "自动识别测量" if getattr(config, "workflow_mode", "Manual") == "Auto" else "手动 ROI 测量"},
            {"项目": "自动基准 Mark", "内容": getattr(config, "auto_reference_label", "")},
            {"项目": "自动待测 Mark", "内容": getattr(config, "auto_target_label", "")},
            {"项目": "像素尺寸X(μm/px)", "内容": config.pixel_size_x_um},
            {"项目": "像素尺寸Y(μm/px)", "内容": config.pixel_size_y_um},
            {"项目": "Rz分布方向", "内容": config.rz_layout},
            {"项目": "Rz单位", "内容": "μrad"},
            {"项目": "Mark间距L(μm)", "内容": config.rz_distance_l_um},
        ]
    info_df = pd.DataFrame(info_rows).round(3)

    with TemporaryDirectory() as tmp_dir:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            info_df.to_excel(writer, index=False, sheet_name="基础信息")
            summary_df.to_excel(writer, index=False, sheet_name="结果汇总")
            detail_df.to_excel(writer, index=False, sheet_name="识别明细")

            for sheet_name in ("基础信息", "结果汇总", "识别明细"):
                ws = writer.book[sheet_name]
                _style_sheet(ws, fail_columns=["结果", "判定", "提示"])

            if mark_images:
                ws = writer.book.create_sheet("Mark图片")
                ws["A1"] = "标记"
                ws["B1"] = "层"
                ws["C1"] = "ROI截图"
                ws["D1"] = "说明"
                _style_sheet(ws)
                for idx, item in enumerate(mark_images, start=2):
                    ws.cell(idx, 1, item.get("mark_id", ""))
                    ws.cell(idx, 2, item.get("layer", ""))
                    ws.cell(idx, 4, item.get("note", ""))
                    image_path = item.get("path")
                    if image_path and Path(image_path).exists():
                        img = XlsxImage(str(image_path))
                        img.width = min(img.width, 260)
                        img.height = min(img.height, 180)
                        ws.add_image(img, f"C{idx}")
                        ws.row_dimensions[idx].height = 140
                ws.column_dimensions["C"].width = 38
